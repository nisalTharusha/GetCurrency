from selenium import webdriver
import time
import os
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.devtools.v98 import browser
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import mysql.connector
import pyodbc
from datetime import datetime


class WebAutomation():

    def __init__(self):
        self.driver = webdriver.Chrome(executable_path=r"E:\Selenium\getcurrecncy\chromedriver.exe")
        self.driver.maximize_window()
        self.wrk="E:\Selenium\getcurrecncy\Data\FX1.xlsx"
        self.wrkbk = openpyxl.load_workbook(self.wrk)

        server = "LAPTOP-MSQ2U88V\SQLEXPRESS"
        Database = "pythontest"
        tcon = "yes"
        uname = "Nisal"
        pword = "newsrilanka"

        self.conx_string = pyodbc.connect(driver='{SQL Server}', host=server, database=Database,
                                     trusted_connection=tcon, user=uname, password=pword)
        self.mycursor = self.conx_string.cursor()

    def Navigating(self):
        driver = self.driver

        try:

            driver.get("https://www.oanda.com/currency-converter/en/?from=AED&to=USD&amount=1")
            time.sleep(4)

            return 'pass'

        except:

               return 'fail'

    def excel(self):
        wrkbk = self.wrkbk
        driver = self.driver
        wrk =self.wrk
        mycursor  =self.mycursor
        conx_string =self.conx_string

        try:
            sh= wrkbk.active
            for i in range(2, sh.max_row + 1):
                cell_obj = sh.cell(row=i, column=1)
                print(cell_obj.value)
                New_link = "https://www.oanda.com/currency-converter/en/?from="+cell_obj.value+"&to=USD&amount=1"
                print(New_link)
                driver.get(New_link)
                element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//body/div[@id='scroll-wrap']/main[1]/div[2]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/button[1]/span[1]/*[1]")))
                currecncy_value = WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.CSS_SELECTOR,"div.block.content-cards-lite-row.card-currency_converter_row.background--white:nth-child(2) div.block__container div.content-cards.content-cards--equal-height div.content-card-lite.content-cards__card.card-main:nth-child(3) div.content-cards__card__inner.background--transparent.border--dove-5.content-cards__card--rounded div.card-content:nth-child(1) div.MuiGrid-root.MuiGrid-container.MuiGrid-justify-xs-space-between:nth-child(2) div.MuiGrid-root.MuiGrid-container.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-sm-true:nth-child(3) div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12:nth-child(2) div.MuiFormControl-root.MuiTextField-root.cc25.MuiFormControl-marginNormal.MuiFormControl-fullWidth:nth-child(1) div.MuiInputBase-root.MuiFilledInput-root.MuiFilledInput-underline.MuiInputBase-fullWidth.MuiInputBase-formControl > input.MuiInputBase-input.MuiFilledInput-input"))).get_attribute('value')

                print("here =" +currecncy_value)

                #get text from the


                valueCel = sh.cell(row=i,column=2)
                valueCel.value =currecncy_value
                wrkbk.save(wrk)

                #save to DB
                now=datetime.now() # current date and time
                Timz = now.strftime("%H:%M:%S")

                code = "insert into [Conversion] values('"+Timz+"','"+cell_obj.value+"','"+currecncy_value+"');"
                print(code)

                mycursor.execute(code)
                conx_string.commit()
                print(mycursor.rowcount, "record inserted.")







            return 'pass'

        except:


            return 'fail'