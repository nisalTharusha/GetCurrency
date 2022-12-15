import pyodbc
import openpyxl
from datetime import datetime
from openpyxl import Workbook

wrk = Workbook(r'E:\Selenium\getcurrecncy\Data\Add.xlsx')
wrkbk =openpyxl.load_workbook(wrk)
Sh = wrkbk.active

for i in range(2,Sh.max_row  + 1):
    Exchange_rate =Sh.cell(row=i,column=2)
    print(Exchange_rate)


